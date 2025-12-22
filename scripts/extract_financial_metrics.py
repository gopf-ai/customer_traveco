#!/usr/bin/env python3
"""
Extract financial metrics from yearly Excel files (2022-2025).

Extracts three key metrics:
- Total Revenue (Sachkonto 35060000 + 35070000)
- Personnel Costs (Sachkonto 0151)
- External Driver Costs (Sachkonto 62802000)

Output: Long format CSV with columns (year, month, metric, value)
"""

import os
import re
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook


# Configuration
YEARS = [2022, 2023, 2024, 2025]
BASE_PATH = Path(__file__).parent.parent / "data" / "raw"
OUTPUT_PATH = Path(__file__).parent.parent / "data" / "processed" / "financial_metrics_overview.csv"

# Account mapping - with fallbacks for different years
# Note: Chart of accounts changed between 2022 and 2023
# - 2022 uses different account structure
# - Leading zeros may be stripped in Excel
ACCOUNTS = {
    'total_revenue': {
        'default': ['35060000', '35070000'],  # Sum of both (2023+)
        # 2022 used different account structure before the accounting system change
        'fallback_2022': ['35010000'],  # Dienstleistungsumsatz (Jan 2022 = -12,518,300.87)
    },
    'personnel_costs': {
        'default': ['0151', '151'],  # Leading zero may be stripped
    },
    'external_driver_costs': {
        'default': ['62802000'],
        # In 2022, LKW wasn't split out until July - use total Ausgangsfrachten account
        # Note: This includes both Bahn (rail) and LKW (truck) for Jan-Jun 2022
        'fallback_2022': ['0628000'],  # Total external freight (summary line)
    },
    'total_betriebsertrag': {
        'default': ['0140'],  # Total Betriebsertrag (Total Operating Revenue)
    },
    'ebt': {
        'default': ['0110'],  # EBT (Earnings Before Tax)
    }
}

# Month columns mapping (C=January, N=December)
MONTH_COLUMNS = {
    1: 'C', 2: 'D', 3: 'E', 4: 'F', 5: 'G', 6: 'H',
    7: 'I', 8: 'J', 9: 'K', 10: 'L', 11: 'M', 12: 'N'
}

# Reference values for validation (Jan 2024)
REFERENCE_VALUES = {
    'total_revenue': -13595064.09,
    'personnel_costs': 5776973.95,
    'external_driver_costs': 1916937.89,
    'total_betriebsertrag': -13803174.46,  # SK 0140, Jan 2024
    'ebt': 329514.32  # SK 0110, Jan 2024
}


def parse_german_number(value) -> Optional[float]:
    """
    Parse German number format to float.

    German format: "5.776.973,95" or "-12.853.254,70"
    Also handles: "5,776,973.95" (CSV export format with quotes)
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    # Convert to string and strip whitespace/quotes
    s = str(value).strip().strip('"').strip("'")

    if not s or s == '':
        return None

    # Determine format by analyzing separators
    # If we have comma before period, it's German format (1.234,56)
    # If we have period before comma, it's English format (1,234.56)

    period_pos = s.rfind('.')
    comma_pos = s.rfind(',')

    if comma_pos > period_pos:
        # German format: periods are thousands, comma is decimal
        # "1.234.567,89" -> "1234567.89"
        s = s.replace('.', '')  # Remove thousands separator
        s = s.replace(',', '.')  # Convert decimal separator
    elif period_pos > comma_pos and comma_pos != -1:
        # English/CSV format: commas are thousands, period is decimal
        # "1,234,567.89" -> "1234567.89"
        s = s.replace(',', '')  # Remove thousands separator
    else:
        # Only one separator or none - check if it could be thousands
        # e.g., "1,234" could be German 1.234 or English 1,234
        if comma_pos != -1 and len(s) - comma_pos - 1 == 2:
            # Exactly 2 digits after comma - likely German decimal
            s = s.replace(',', '.')
        elif comma_pos != -1:
            # More digits after comma - likely English thousands
            s = s.replace(',', '')

    try:
        return float(s)
    except ValueError:
        print(f"Warning: Could not parse number: '{value}'")
        return None


def get_excel_path(year: int) -> Path:
    """Get the path to the Excel file for a given year."""
    return BASE_PATH / str(year) / f"{year} Finanzen" / f"{year}.xlsx"


def get_account_codes(metric_name: str, year: int) -> list[str]:
    """Get the account codes to use for a metric in a given year."""
    config = ACCOUNTS[metric_name]

    # Check for year-specific fallback
    fallback_key = f'fallback_{year}'
    if fallback_key in config:
        return config[fallback_key]

    return config.get('default', [])


def extract_year_data(year: int) -> list[dict]:
    """
    Extract financial metrics for a given year.

    Returns list of dicts with keys: year, month, metric, value
    """
    excel_path = get_excel_path(year)

    if not excel_path.exists():
        print(f"Warning: File not found: {excel_path}")
        return []

    print(f"Processing {year}...")

    # Load workbook
    wb = load_workbook(excel_path, data_only=True)

    # Get the correct sheet
    sheet_name = "Allgemeiner Report"
    if sheet_name not in wb.sheetnames:
        # Try to find a similar sheet name
        for name in wb.sheetnames:
            if "Report" in name or "Allgemein" in name:
                sheet_name = name
                break
        else:
            print(f"Warning: Sheet 'Allgemeiner Report' not found in {excel_path}")
            print(f"Available sheets: {wb.sheetnames}")
            return []

    ws = wb[sheet_name]

    # Build a mapping of Sachkonto -> row data
    sachkonto_data = {}

    for row in ws.iter_rows(min_row=4):  # Data starts at row 4
        sachkonto = str(row[0].value).strip() if row[0].value else ""
        if sachkonto:
            sachkonto_data[sachkonto] = row

    # Determine how many months to extract (2025 has 11 months through November)
    max_month = 11 if year == 2025 else 12

    results = []
    warned_accounts = set()

    for metric_name in ACCOUNTS.keys():
        account_codes = get_account_codes(metric_name, year)

        for month in range(1, max_month + 1):
            col_letter = MONTH_COLUMNS[month]
            col_idx = ord(col_letter) - ord('A')  # Convert letter to 0-based index

            total_value = 0.0
            found_any = False

            for account_code in account_codes:
                if account_code in sachkonto_data:
                    row = sachkonto_data[account_code]
                    cell_value = row[col_idx].value if col_idx < len(row) else None
                    parsed_value = parse_german_number(cell_value)

                    if parsed_value is not None:
                        total_value += parsed_value
                        found_any = True
                else:
                    # Only warn once per account per year
                    warn_key = f"{account_code}_{year}"
                    if warn_key not in warned_accounts:
                        print(f"  Note: Account {account_code} not found for {year} ({metric_name})")
                        warned_accounts.add(warn_key)

            if found_any:
                results.append({
                    'year': year,
                    'month': month,
                    'metric': metric_name,
                    'value': total_value
                })

    wb.close()
    return results


def validate_reference_values(df: pd.DataFrame) -> bool:
    """Validate Jan 2024 values against reference."""
    print("\n=== Validation against reference values (Jan 2024) ===")

    jan_2024 = df[(df['year'] == 2024) & (df['month'] == 1)]

    all_valid = True
    for metric, expected in REFERENCE_VALUES.items():
        actual_row = jan_2024[jan_2024['metric'] == metric]
        if actual_row.empty:
            print(f"  {metric}: MISSING!")
            all_valid = False
        else:
            actual = actual_row['value'].values[0]
            diff = abs(actual - expected)
            diff_pct = abs(diff / expected * 100) if expected != 0 else 0

            if diff_pct < 0.01:  # Less than 0.01% difference
                print(f"  {metric}: {actual:,.2f} (expected {expected:,.2f}) - MATCH")
            else:
                print(f"  {metric}: {actual:,.2f} (expected {expected:,.2f}) - MISMATCH ({diff_pct:.2f}%)")
                all_valid = False

    return all_valid


def main():
    """Main extraction routine."""
    print("=" * 60)
    print("Financial Metrics Extraction")
    print("=" * 60)

    all_results = []

    for year in YEARS:
        year_data = extract_year_data(year)
        all_results.extend(year_data)
        print(f"  Extracted {len(year_data)} records for {year}")

    # Convert to DataFrame
    df = pd.DataFrame(all_results)

    # Sort by year, month, metric
    df = df.sort_values(['year', 'month', 'metric']).reset_index(drop=True)

    # Validate
    validation_passed = validate_reference_values(df)

    # Summary statistics
    print("\n=== Summary ===")
    print(f"Total records: {len(df)}")
    print(f"Years covered: {df['year'].min()} - {df['year'].max()}")
    print(f"Metrics: {df['metric'].unique().tolist()}")

    # Show sample data
    print("\n=== Sample data (first 10 rows) ===")
    print(df.head(10).to_string(index=False))

    # Save to CSV
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUTPUT_PATH, index=False)
    print(f"\n=== Output saved to: {OUTPUT_PATH} ===")

    if not validation_passed:
        print("\nWARNING: Validation failed - please check the data!")
    else:
        print("\nValidation passed!")

    return df


if __name__ == "__main__":
    main()
