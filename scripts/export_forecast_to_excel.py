#!/usr/bin/env python3
"""
Export forecast data to Excel matching the dashboard "Monatliche Details" format.

Columns (matching forecast_dashboard_2025_2026.html):
- Jahr (Year)
- Monat (Month - German abbreviations)
- Typ (Ist/Prognose)
- Aufträge (Orders)
- Umsatz (Op.) (Operational Revenue)
- Total Betr.ertrag (SK 0140)
- Betriebsertrag (SK 3506+3507)
- Personalaufw. (SK 0151)
- Ausg.frachten (SK 6280)

Note: EBT (SK 0110) is excluded as per requirements.
"""

import warnings
warnings.filterwarnings('ignore')

import pandas as pd
import numpy as np
from pathlib import Path

# Paths
BASE_PATH = Path(__file__).parent.parent
DATA_PATH = BASE_PATH / "data" / "processed"
OUTPUT_PATH = BASE_PATH / "results" / "Traveco_Forecast_2022_2026.xlsx"

# German month abbreviations (matching dashboard)
MONTH_ABBREV_DE = {
    1: 'Jan', 2: 'Feb', 3: 'Mär', 4: 'Apr',
    5: 'Mai', 6: 'Jun', 7: 'Jul', 8: 'Aug',
    9: 'Sep', 10: 'Okt', 11: 'Nov', 12: 'Dez'
}


def load_operational_data():
    """Load operational data (historical 2022-2024 + forecasts 2025-2026)."""
    # Historical data (2022-2024)
    hist_df = pd.read_csv(DATA_PATH / "monthly_aggregated_full_company.csv")
    hist_df['date'] = pd.to_datetime(hist_df['date'])
    hist_df = hist_df[hist_df['date'].dt.year <= 2024]
    hist_df['source'] = 'actual'

    # 2025-2026 data (actuals + forecasts)
    forecast_df = pd.read_csv(DATA_PATH / "combined_forecast_2025_2026.csv")
    forecast_df['date'] = pd.to_datetime(forecast_df['date'])

    # Combine
    hist_subset = hist_df[['date', 'total_orders', 'revenue_total', 'source']].copy()
    forecast_subset = forecast_df[['date', 'total_orders', 'revenue_total', 'source']].copy()

    combined = pd.concat([hist_subset, forecast_subset], ignore_index=True)
    combined = combined.sort_values('date').reset_index(drop=True)

    return combined


def load_financial_data():
    """Load financial data (actuals 2022-Nov 2025 + forecasts Dec 2025-Dec 2026)."""
    # Actuals (2022 - Nov 2025)
    actuals_df = pd.read_csv(DATA_PATH / "financial_metrics_overview.csv")
    actuals_df['date'] = pd.to_datetime(
        actuals_df['year'].astype(str) + '-' +
        actuals_df['month'].astype(str).str.zfill(2) + '-01'
    )
    actuals_df['source'] = 'actual'

    # Forecasts (Dec 2025 - Dec 2026)
    forecasts_df = pd.read_csv(DATA_PATH / "financial_metrics_forecasts.csv")
    forecasts_df['date'] = pd.to_datetime(forecasts_df['ds'])
    forecasts_df = forecasts_df.rename(columns={'prediction': 'value'})
    forecasts_df['source'] = 'forecast'

    return actuals_df, forecasts_df


def build_financial_time_series(metric_name, actuals_df, forecasts_df):
    """Build time series for a financial metric."""
    actual_data = actuals_df[actuals_df['metric'] == metric_name][['date', 'value', 'source']].copy()
    forecast_data = forecasts_df[forecasts_df['metric'] == metric_name][['date', 'value', 'source']].copy()
    combined = pd.concat([actual_data, forecast_data], ignore_index=True)
    combined = combined.sort_values('date').reset_index(drop=True)
    return combined


def build_data_table(operational_df, fin_actuals, fin_forecasts):
    """Build comprehensive data table matching dashboard format (excluding EBT)."""
    # Create date range from 2022-01 to 2026-12
    dates = pd.date_range('2022-01-01', '2026-12-01', freq='MS')
    table_df = pd.DataFrame({'date': dates})
    table_df['Jahr'] = table_df['date'].dt.year
    table_df['Monat'] = table_df['date'].dt.month.map(MONTH_ABBREV_DE)

    # Add operational metrics (total_orders, revenue_total)
    for metric in ['total_orders', 'revenue_total']:
        ts = operational_df[['date', metric]].copy()
        table_df = table_df.merge(ts, on='date', how='left')

    # Add financial metrics (excluding EBT)
    for metric in ['total_betriebsertrag', 'total_revenue', 'personnel_costs', 'external_driver_costs']:
        ts = build_financial_time_series(metric, fin_actuals, fin_forecasts)
        # Invert revenue metrics (negative in accounting)
        if metric in ['total_revenue', 'total_betriebsertrag']:
            ts['value'] = -ts['value']
        ts = ts.rename(columns={'value': metric})
        table_df = table_df.merge(ts[['date', metric]], on='date', how='left')

    # Add Typ column (Ist/Prognose)
    table_df['Typ'] = table_df['date'].apply(
        lambda d: 'Prognose' if d >= pd.Timestamp('2025-12-01') else 'Ist'
    )

    return table_df


def export_to_excel(table_df, output_file):
    """Export the data table to Excel with proper formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Select and rename columns for output (matching dashboard)
    output_df = table_df[[
        'Jahr', 'Monat', 'Typ',
        'total_orders', 'revenue_total',
        'total_betriebsertrag', 'total_revenue',
        'personnel_costs', 'external_driver_costs'
    ]].copy()

    # Rename columns to match dashboard exactly
    output_df.columns = [
        'Jahr', 'Monat', 'Typ',
        'Aufträge', 'Umsatz (Op.)',
        'Total Betr.ertrag (SK 0140)', 'Betriebsertrag (SK 3506+3507)',
        'Personalaufw. (SK 0151)', 'Ausg.frachten (SK 6280)'
    ]

    # Round numeric columns to integers
    numeric_cols = [
        'Aufträge', 'Umsatz (Op.)',
        'Total Betr.ertrag (SK 0140)', 'Betriebsertrag (SK 3506+3507)',
        'Personalaufw. (SK 0151)', 'Ausg.frachten (SK 6280)'
    ]
    for col in numeric_cols:
        output_df[col] = output_df[col].round(0).astype('Int64')

    # Export to Excel with formatting
    output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        output_df.to_excel(writer, index=False, sheet_name='Forecast 2022-2026')

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Forecast 2022-2026']

        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='00852A', end_color='00852A', fill_type='solid')  # Traveco green
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        forecast_fill = PatternFill(start_color='FFF8E6', end_color='FFF8E6', fill_type='solid')  # Light yellow

        number_alignment = Alignment(horizontal='right')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Style header row
        for col_idx, col in enumerate(output_df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Style data rows
        for row_idx in range(2, len(output_df) + 2):
            is_forecast = worksheet.cell(row=row_idx, column=3).value == 'Prognose'

            for col_idx in range(1, len(output_df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

                # Apply forecast highlight
                if is_forecast:
                    cell.fill = forecast_fill

                # Right-align numeric columns (4 onwards)
                if col_idx >= 4:
                    cell.alignment = number_alignment
                    # Format as number with thousand separator
                    cell.number_format = '#,##0'

        # Set column widths
        column_widths = {
            'A': 8,   # Jahr
            'B': 8,   # Monat
            'C': 12,  # Typ
            'D': 14,  # Aufträge
            'E': 16,  # Umsatz (Op.)
            'F': 26,  # Total Betr.ertrag
            'G': 28,  # Betriebsertrag
            'H': 20,  # Personalaufw.
            'I': 20,  # Ausg.frachten
        }

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        # Set row height for header
        worksheet.row_dimensions[1].height = 30

        # Freeze header row
        worksheet.freeze_panes = 'A2'

    print(f"  Excel file saved: {output_file}")
    return output_file


def main():
    """Main function."""
    print("=" * 70)
    print("Exporting Traveco Forecast to Excel (Dashboard Format)")
    print("=" * 70)

    print("\n1. Loading operational data...")
    operational_df = load_operational_data()
    print(f"   Loaded {len(operational_df)} operational records")

    print("\n2. Loading financial data...")
    fin_actuals, fin_forecasts = load_financial_data()
    print(f"   Loaded {len(fin_actuals)} actuals, {len(fin_forecasts)} forecasts")

    print("\n3. Building data table...")
    table_df = build_data_table(operational_df, fin_actuals, fin_forecasts)
    print(f"   Created table with {len(table_df)} rows (Jan 2022 - Dec 2026)")

    print("\n4. Exporting to Excel...")
    export_to_excel(table_df, OUTPUT_PATH)

    print("\n" + "=" * 70)
    print("EXPORT COMPLETE!")
    print("=" * 70)
    print(f"\nOutput file: {OUTPUT_PATH}")
    print("\nColumns:")
    print("  - Jahr, Monat, Typ")
    print("  - Aufträge, Umsatz (Op.)")
    print("  - Total Betr.ertrag (SK 0140), Betriebsertrag (SK 3506+3507)")
    print("  - Personalaufw. (SK 0151), Ausg.frachten (SK 6280)")
    print("\nNote: EBT column excluded as requested.")

    return OUTPUT_PATH


if __name__ == "__main__":
    main()
